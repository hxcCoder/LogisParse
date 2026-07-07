# app/services/excel_service.py
"""
Servicio de generación de archivos Excel (.xlsx) con los datos extraídos.
Organiza la información en columnas limpias siguiendo el formato SII chileno.

Dependencias:
    pip install openpyxl
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── CONSTANTES DE FORMATO ──

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# ── COLUMNAS DEL REPORTE ──

COLUMNS = [
    ("RUT Emisor", "rut_emisor", 18),
    ("RUT Receptor", "rut_receptor", 18),
    ("Folio SII / Guía", "folio_sii", 20),
    ("Fecha Emisión", "fecha_emision", 14),
    ("Monto Total ($)", "monto_total", 15),
    ("Origen", "origen", 20),
    ("Destino", "destino", 20),
    ("Patente Camión", "patente_camion", 15),
    ("Chofer", "chofer", 30),
    ("Fecha Despacho", "fecha_despacho", 14),
    ("N° Guía Despacho", "numero_guia", 18),
    ("Adaptador usado", "adapter_used", 25),
    ("Confianza (%)", "confidence_score", 12),
    ("Observaciones", "observaciones", 40),
]


class ExcelService:
    """
    Genera archivos Excel formateados a partir de los datos extraídos.
    """

    def __init__(self) -> None:
        self.workbook: Workbook | None = None
        self.sheet: Any = None

    def create_workbook(
        self,
        extracted_data_list: list[dict[str, Any]],
        sheet_name: str = "Documentos Procesados",
    ) -> Workbook:
        """
        Crea un libro de Excel con los datos extraídos.

        Args:
            extracted_data_list: Lista de diccionarios con los datos extraídos.
            sheet_name: Nombre de la hoja de cálculo.

        Returns:
            Libro de Excel (Workbook) listo para guardar.
        """
        if not extracted_data_list:
            raise ValueError("La lista de datos extraídos no puede estar vacía.")

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = sheet_name

        self._write_headers()
        self._write_data(extracted_data_list)
        self._auto_adjust_columns()
        self._freeze_header()

        logger.info(
            "Excel generado exitosamente con %d registros.",
            len(extracted_data_list),
        )
        return self.workbook

    def save_to_bytes(self) -> bytes:
        """
        Guarda el workbook en memoria como bytes.

        Returns:
            Bytes del archivo Excel (.xlsx).

        Raises:
            RuntimeError: Si no se ha creado un workbook primero.
        """
        if self.workbook is None:
            raise RuntimeError(
                "No hay un workbook creado. Llama a create_workbook() primero."
            )

        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        logger.info("Excel guardado en memoria (%d bytes).", output.getbuffer().nbytes)
        return output.getvalue()

    def save_to_file(self, filepath: str | Path) -> str:
        """
        Guarda el workbook en disco.

        Args:
            filepath: Ruta donde guardar el archivo.

        Returns:
            Ruta absoluta del archivo guardado.

        Raises:
            RuntimeError: Si no se ha creado un workbook.
        """
        if self.workbook is None:
            raise RuntimeError(
                "No hay un workbook creado. Llama a create_workbook() primero."
            )

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)

        self.workbook.save(str(path))
        logger.info("Excel guardado en: %s", path.absolute())
        return str(path.absolute())

    # ── MÉTODOS PRIVADOS ──

    def _write_headers(self) -> None:
        """Escribe la fila de encabezados con formato profesional."""
        if self.sheet is None:
            return

        for col_idx, (label, _, _) in enumerate(COLUMNS, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Altura de fila de encabezado
        self.sheet.row_dimensions[1].height = 30

    def _write_data(self, data_list: list[dict[str, Any]]) -> None:
        """Escribe las filas de datos extraídos."""
        if self.sheet is None:
            return

        for row_idx, data in enumerate(data_list, start=2):
            for col_idx, (_, field_name, _) in enumerate(COLUMNS, start=1):
                value = data.get(field_name, "")

                # Formatear confidence_score como porcentaje
                if field_name == "confidence_score" and isinstance(value, (int, float)):
                    value = f"{value:.1f}%"

                # Formatear monto_total con separador de miles
                if field_name == "monto_total" and value:
                    try:
                        value = f"${int(float(str(value))):,}"
                    except (ValueError, TypeError):
                        pass

                cell = self.sheet.cell(row=row_idx, column=col_idx, value=value or "—")
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGNMENT
                cell.border = THIN_BORDER

            # Alternar color de fila para mejor legibilidad
            if row_idx % 2 == 0:
                for col_idx in range(1, len(COLUMNS) + 1):
                    self.sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"
                    )

    def _auto_adjust_columns(self) -> None:
        """Ajusta el ancho de las columnas según el contenido."""
        if self.sheet is None:
            return

        for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
            self.sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _freeze_header(self) -> None:
        """Congela la fila de encabezados para facilitar el scroll."""
        if self.sheet is not None:
            self.sheet.freeze_panes = "A2"


# ── FUNCIÓN DE CONVENIENCIA ──

def generate_excel_report(
    data: list[dict[str, Any]],
    output_path: str | None = None,
) -> bytes:
    """
    Genera un reporte Excel a partir de una lista de datos extraídos.

    Args:
        data: Lista de diccionarios con datos extraídos.
        output_path: Ruta opcional para guardar el archivo en disco.

    Returns:
        Bytes del archivo Excel.

    Raises:
        ValueError: Si no hay datos para generar el reporte.
    """
    if not data:
        raise ValueError("No hay datos para generar el reporte Excel.")

    service = ExcelService()
    service.create_workbook(data)

    if output_path:
        service.save_to_file(output_path)
        logger.info("Reporte Excel exportado a: %s", output_path)

    return service.save_to_bytes()